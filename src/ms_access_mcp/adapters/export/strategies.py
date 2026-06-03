"""Export strategy pattern — pluggable format converters.

Each strategy knows how to write query results to a specific file format.
Strategies receive callables from the adapter so they are decoupled from
both WinComAdapter and OdbcAdapter internals.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Protocol


@dataclass
class ExportContext:
    """Everything a strategy needs to write data to a file.

    Attributes:
        sql: Raw SELECT query to execute.
        file_path: Destination file path.
        options: Format-specific options such as ``delimiter``, ``header``,
            ``encoding``, ``pretty``, ``sheet_name``.
        execute_query: Bound ``adapter.execute_query(sql)`` callable.
            Returns ``{"success": bool, "rows": list[dict], "columns": list[str]}``.
        execute_raw: Bound callable that runs arbitrary SQL through the
            Access engine (IISAM path, DDL, etc.).
            Returns int (rows affected) on success, raises on error.
    """

    sql: str
    file_path: str
    options: dict[str, Any] = field(default_factory=dict)
    execute_query: Callable[[str], dict] | None = None
    execute_raw: Callable[[str], int] | None = None


class ExportStrategy(Protocol):
    """Protocol for a single export format."""

    format_name: str

    def export(self, context: ExportContext) -> dict[str, Any]:
        """Export data to *context.file_path* in this strategy's format.

        Returns:
            A dict with ``success`` (bool), ``rows_exported`` (int),
            ``file_path`` (str), and optionally ``error`` (str).
        """
        ...


# =============================================================================
# CSV  —  Text IISAM fast path  +  csv.DictWriter fallback
# =============================================================================

_CODEPAGE_MAP: dict[str, int] = {
    "utf-8": 65001,
    "utf-16": 1200,
    "latin-1": 28591,
    "windows-1252": 1252,
    "cp1252": 1252,
    "shift-jis": 932,
}


class CsvStrategy:
    """Export rows to CSV.

    IISAM path (fast): ``INSERT INTO [Text;…] SELECT …`` via the Access engine.
    Fallback: fetch rows in Python and write via ``csv.DictWriter``.
    """

    format_name = "csv"

    def export(self, context: ExportContext) -> dict[str, Any]:
        import csv

        delimiter = context.options.get("delimiter", ",")
        header = context.options.get("header", True)
        encoding = context.options.get("encoding", "utf-8")

        # --- IISAM fast path (default delimiter + known codepage) -----------
        if delimiter == ",":
            codepage = _CODEPAGE_MAP.get(encoding)
            if codepage is not None and context.execute_raw is not None:
                result = self._try_iisam(context, header, encoding, codepage)
                if result is not None:
                    return result

        # --- Fallback: fetch rows + csv.DictWriter --------------------------
        return self._fallback(context, delimiter, header, encoding)

    # -- private helpers -----------------------------------------------------

    def _try_iisam(
        self,
        context: ExportContext,
        header: bool,
        encoding: str,
        codepage: int,
    ) -> dict[str, Any] | None:
        p = Path(context.file_path)
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            conn_str = (
                f"Text;FMT=Delimited;HDR={'YES' if header else 'NO'};"
                f"CharacterSet={codepage};DATABASE={p.parent.absolute()}"
            )
            insert_sql = f"INSERT INTO [{conn_str}].[{p.name}] {context.sql}"
            affected = context.execute_raw(insert_sql)
            return {
                "success": True,
                "rows_exported": affected,
                "file_path": context.file_path,
            }
        except Exception:
            return None

    def _fallback(
        self,
        context: ExportContext,
        delimiter: str,
        header: bool,
        encoding: str,
    ) -> dict[str, Any]:
        import csv

        if context.execute_query is None:
            return {"success": False, "error": "execute_query not available for CSV fallback"}

        result = context.execute_query(context.sql)
        if not result.get("success"):
            return {"success": False, "error": result.get("error", "Query failed")}

        rows: list[dict] = result.get("rows", [])
        columns: list[str] = result.get("columns", [])

        try:
            Path(context.file_path).parent.mkdir(parents=True, exist_ok=True)
            with open(context.file_path, "w", newline="", encoding=encoding) as f:
                writer = csv.DictWriter(f, fieldnames=columns, delimiter=delimiter)
                if header:
                    writer.writeheader()
                writer.writerows(rows)
            return {
                "success": True,
                "rows_exported": len(rows),
                "file_path": context.file_path,
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


# =============================================================================
# JSON  —  direct json.dump (no IISAM available)
# =============================================================================


class JsonStrategy:
    """Export rows to a JSON array of objects.

    Always uses Python-side ``json.dump`` — no Access engine IISAM exists for JSON.
    """

    format_name = "json"

    def export(self, context: ExportContext) -> dict[str, Any]:
        import json

        if context.execute_query is None:
            return {"success": False, "error": "execute_query not available for JSON export"}

        pretty = context.options.get("pretty", False)

        result = context.execute_query(context.sql)
        if not result.get("success"):
            return {"success": False, "error": result.get("error", "Query failed")}

        rows: list[dict] = result.get("rows", [])

        try:
            Path(context.file_path).parent.mkdir(parents=True, exist_ok=True)
            with open(context.file_path, "w", encoding="utf-8") as f:
                json.dump(rows, f, indent=2 if pretty else None)
            return {
                "success": True,
                "rows_exported": len(rows),
                "file_path": context.file_path,
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


# =============================================================================
# Excel  —  Excel 12.0 IISAM fast path  +  openpyxl fallback
# =============================================================================


class ExcelStrategy:
    """Export rows to Excel (.xlsx).

    IISAM path (fast): ``INSERT INTO [Excel 12.0;…] SELECT …`` via the Access
    engine.  Requires the Microsoft ACE or JET ODBC driver (Windows).

    Fallback: fetch rows in Python and write via ``openpyxl``.
    Works cross-platform (WSL, macOS, Linux) when ``openpyxl`` is installed.
    """

    format_name = "excel"

    def export(self, context: ExportContext) -> dict[str, Any]:
        # --- IISAM fast path (Windows + ACE/JET driver) ---------------------
        if context.execute_raw is not None:
            result = self._try_iisam(context)
            if result is not None:
                return result

        # --- Fallback: openpyxl ---------------------------------------------
        return self._fallback(context)

    # -- private helpers -----------------------------------------------------

    def _try_iisam(self, context: ExportContext) -> dict[str, Any] | None:
        p = Path(context.file_path)
        sheet_name = context.options.get("sheet_name", "Sheet1")
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            conn_str = f"Excel 12.0;HDR=YES;DATABASE={p.absolute()}"
            insert_sql = (
                f"INSERT INTO [{conn_str}].[{sheet_name}$] {context.sql}"
            )
            affected = context.execute_raw(insert_sql)
            return {
                "success": True,
                "rows_exported": affected,
                "file_path": context.file_path,
            }
        except Exception:
            return None

    def _fallback(self, context: ExportContext) -> dict[str, Any]:
        if context.execute_query is None:
            return {"success": False, "error": "execute_query not available for Excel fallback"}

        result = context.execute_query(context.sql)
        if not result.get("success"):
            return {"success": False, "error": result.get("error", "Query failed")}

        rows: list[dict] = result.get("rows", [])
        columns: list[str] = result.get("columns", [])
        sheet_name = context.options.get("sheet_name", "Sheet1")

        try:
            Path(context.file_path).parent.mkdir(parents=True, exist_ok=True)
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel limit: 31 chars

            # Header
            ws.append(columns)
            # Data
            for row in rows:
                ws.append([row.get(col, "") for col in columns])

            # Auto-adjust column widths
            _auto_width(ws, columns, rows)

            wb.save(context.file_path)
            return {
                "success": True,
                "rows_exported": len(rows),
                "file_path": context.file_path,
            }
        except ImportError:
            return {
                "success": False,
                "error": "openpyxl is not installed. Run: pip install ms-access-mcp[excel]",
            }
        except Exception as e:
            return {"success": False, "error": str(e)}


def _auto_width(ws: Any, columns: list[str], rows: list[dict]) -> None:
    """Set reasonable column widths based on header + data length."""
    from openpyxl.utils import get_column_letter

    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row in rows:
            val = row.get(col_name, "")
            max_len = max(max_len, len(str(val)))
        # Cap at 50, minimum 8
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 8), 50)


# =============================================================================
# Selector  —  registry of available strategies
# =============================================================================


class ExportStrategySelector:
    """Registry of ``ExportStrategy`` implementations.

    Usage::

        selector = ExportStrategySelector()
        strategy = selector.get("csv")
        result = strategy.export(context)
    """

    def __init__(self) -> None:
        self._strategies: dict[str, ExportStrategy] = {}
        self._register(CsvStrategy())
        self._register(JsonStrategy())
        self._register(ExcelStrategy())

    def _register(self, strategy: ExportStrategy) -> None:
        self._strategies[strategy.format_name] = strategy

    def register(self, strategy: ExportStrategy) -> None:
        """Register a custom strategy (overrides an existing one if same
        ``format_name``)."""
        self._strategies[strategy.format_name] = strategy

    def get(self, format_name: str) -> ExportStrategy:
        """Look up a strategy by its format name.

        Raises ``ValueError`` if the format is not registered.
        """
        strategy = self._strategies.get(format_name)
        if strategy is None:
            raise ValueError(
                f"Unsupported export format: {format_name!r}. "
                f"Supported: {', '.join(self.supported_formats())}"
            )
        return strategy

    def supported_formats(self) -> list[str]:
        """Return the list of registered format names."""
        return list(self._strategies.keys())
